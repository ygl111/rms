"""
Excel 处理工具模块
提供生成 Excel 文件的通用功能
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.cell import WriteOnlyCell
from io import BytesIO
from datetime import datetime


def create_excel_file(data, headers, sheet_name="Data Export"):
    """
    创建 Excel 文件
    :param data: 要导出的数据列表，每个元素是一个字典
    :param headers: 表头配置，格式为 [{'key': '字段名', 'title': '显示标题', 'width': 宽度}, ...]
    :param sheet_name: 工作表名称
    :return: Excel 文件的字节流
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")  # 白色粗体字
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色背景
    header_alignment = Alignment(horizontal="center", vertical="center")  # 居中对齐
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header['title'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置列宽
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = header.get('width', 15)
    
    # 写入数据
    for row_idx, item in enumerate(data, 2):  # 从第2行开始（第1行是表头）
        for col_idx, header in enumerate(headers, 1):
            # 获取字段值，支持嵌套字段（如 role.role_name）
            value = get_nested_value(item, header['key'])
            # 清理 Excel 不允许的控制字符
            if isinstance(value, str):
                value = ILLEGAL_CHARACTERS_RE.sub('', value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # 文本对齐
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    # 将工作簿保存到字节流中
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def _sanitize_excel_value(value):
    """统一清理 Excel 单元格的值，确保类型与非法字符处理一致。"""
    if value is None:
        return ""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def create_excel_file_streaming(headers, row_iterable, sheet_name="Data Export"):
    """
    针对大数据量导出的轻量级 Excel 生成器，采用 write_only 模式并逐行写入，降低内存占用。

    :param headers: 表头配置，格式同 create_excel_file
    :param row_iterable: 可迭代对象，每次迭代返回与 headers 顺序一致的序列（list/tuple）
    :param sheet_name: 工作表名称
    :return: Excel 文件的字节流
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_name)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    header_row = []
    for col_idx, header in enumerate(headers, 1):
        cell = WriteOnlyCell(ws, value=header['title'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_row.append(cell)
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = header.get('width', 15)

    ws.append(header_row)

    for row in row_iterable:
        sanitized_row = [_sanitize_excel_value(value) for value in row]
        ws.append(sanitized_row)

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def get_nested_value(obj, key):
    """
    获取嵌套对象的值
    :param obj: 对象（可能是字典或对象实例）
    :param key: 键名，支持点号分隔的嵌套键（如 'role.role_name'）
    :return: 对应的值，如果不存在返回空字符串
    """
    try:
        # 分割键名
        keys = key.split('.')
        value = obj
        
        for k in keys:
            if hasattr(value, k):
                # 如果是对象属性
                value = getattr(value, k)
            elif isinstance(value, dict) and k in value:
                # 如果是字典键
                value = value[k]
            else:
                return ""
        
        # 处理 None 值
        if value is None:
            return ""
        
        # 处理日期时间格式
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
            
        return str(value)
    except:
        return ""


def generate_filename(prefix="导出数据"):
    """
    生成带时间戳的文件名
    :param prefix: 文件名前缀
    :return: 文件名字符串
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"
